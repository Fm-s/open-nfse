#!/usr/bin/env python3
"""
Deterministic extractor for specs/ruleset/.

Reads the official NFS-e Anexo spreadsheets (XLSX) and emits:
  - specs/ruleset/_raw/<slug>.json   lossless per-sheet dumps (provenance backing)
  - specs/ruleset/*.json             clean, typed sidecars derived mechanically

The tabular fiscal data (field layouts, business rules, rejection codes, the
event x event state matrix, IndOp catalog) is extracted ROW-BY-ROW with no
paraphrasing, so every clean row traces back to an exact source cell via `fonte`.
Re-run after the Receita updates any source file.

Usage:  python3 scripts/extract-ruleset.py
"""
from __future__ import annotations
import json
import os
import re
import sys
import unicodedata

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DOC = os.path.join(ROOT, "specs/oficial/documentacao-atual")
OUT = os.path.join(ROOT, "specs/ruleset")
RAW = os.path.join(OUT, "_raw")

A1 = os.path.join(DOC, "anexo_i-sefin_adn-dps_nfse-snnfse-v1-01-20260209.xlsx")
A2 = os.path.join(DOC, "anexo_ii-sefin_adn-pedregevt_evt-snnfse-v1-01-20260122.xlsx")
AC = os.path.join(DOC, "anexo_c-indop_ibscbs-snnfse-v1-01-20260122.xlsx")

# short tokens used in `fonte` strings
SRC = {A1: "AnexoI", A2: "AnexoII", AC: "AnexoC"}


# --------------------------------------------------------------------------- #
# grid loading (merged cells filled so continuation rows inherit their value)
# --------------------------------------------------------------------------- #
def clean(v):
    if v is None:
        return None
    s = str(v)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # collapse runs of spaces/tabs but keep newlines (lists in rule text)
    s = "\n".join(re.sub(r"[ \t]+", " ", ln).strip() for ln in s.split("\n"))
    s = s.strip()
    return s if s != "" else None


def load_grid(path, sheet):
    """Return (max_row, max_col, grid) where grid[r][c] (1-based) is the filled,
    cleaned string value (None if empty). Merged regions are filled with the
    top-left value across every cell so continuation rows inherit it."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    mr, mc = ws.max_row, ws.max_column
    grid = [[None] * (mc + 1) for _ in range(mr + 1)]
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            grid[r][c] = clean(ws.cell(row=r, column=c).value)
    for rng in ws.merged_cells.ranges:
        tl = clean(ws.cell(row=rng.min_row, column=rng.min_col).value)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if grid[r][c] is None:
                    grid[r][c] = tl
    wb.close()
    return mr, mc, grid


def flat(s):
    """collapse all whitespace (incl. newlines) to single spaces — for labels"""
    if s is None:
        return None
    return re.sub(r"\s+", " ", s).strip() or None


def norm(s):
    """accent-insensitive, lowercased, whitespace-collapsed key for name matching"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s.replace("\n", " ")).strip().lower()
    s = s.rstrip(" .:-")
    return s


# raw lossless dump ---------------------------------------------------------- #
def dump_raw(path, sheet, slug):
    mr, mc, grid = load_grid(path, sheet)
    rows = []
    for r in range(1, mr + 1):
        cells = [grid[r][c] for c in range(1, mc + 1)]
        if any(v is not None for v in cells):
            rows.append({"r": r, "cells": cells})
    payload = {
        "source": os.path.basename(path),
        "sheet": sheet,
        "dims": {"rows": mr, "cols": mc},
        "rows": rows,
    }
    write_json(os.path.join(RAW, f"{slug}.json"), payload)
    return mr, mc, grid


def write_json(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")


# --------------------------------------------------------------------------- #
# LEIAUTE (field layout):  # | CAMINHO | CAMPO | ELE | TIPO | OCOR | TAM | DESC | NOTAS
# --------------------------------------------------------------------------- #
def extract_campos(path, sheet, src):
    mr, mc, grid = load_grid(path, sheet)
    out = []
    for r in range(2, mr + 1):  # row 1 = header
        seq, caminho, campo, ele, tipo, ocor, tam, desc, notas = (
            grid[r][c] if c <= mc else None for c in range(1, 10)
        )
        if seq is None and campo is None and caminho is None:
            continue
        out.append({
            "seq": to_int(seq),
            "caminho": caminho,
            "campo": campo,
            "ele": ele,
            "tipo": tipo,
            "ocorrencia": ocor,
            "tamanho": tam,
            "descricao": desc,
            "notas": notas,
            "fonte": f"{src}!{sheet}:r{r}",
        })
    return out


# --------------------------------------------------------------------------- #
# RN (business rules):  # | CAMINHO | CAMPO | REGRA(col3) | (col4) | APLIC(5)
#   | EFEITO(6) | COD(7) | MSG(8) | NIVEL(9) | SEFIN-exec(10) | SEFIN-dec(11)
#   | ADN-exec(12) | ADN-dec(13) | OBS(14).  Data starts row 4 (3-row header).
# --------------------------------------------------------------------------- #
def extract_regras_rn(path, sheet, src, first_data_row=4):
    mr, mc, grid = load_grid(path, sheet)
    out = []
    for r in range(first_data_row, mr + 1):
        g = lambda c: grid[r][c] if c <= mc else None
        seq, caminho, campo = g(1), g(2), g(3)
        regra, aplic, efeito, cod, msg = g(4), g(6), g(7), g(8), g(9)
        nivel = g(10)
        sefin_exec, sefin_dec, adn_exec, adn_dec, obs = g(11), g(12), g(13), g(14), g(15)
        # a real rule has rule text or a rejection code; otherwise it is a
        # structural/group row (covered by the campos layout) -> skip.
        has_rule = (regra not in (None, "-")) or (cod not in (None, "-"))
        if not has_rule:
            continue
        out.append({
            "seq": to_int(seq),
            "caminho": caminho,
            "campo": campo,
            "regra": dash_none(regra),
            "aplic": dash_none(aplic),
            "efeito": dash_none(efeito),
            "codErro": dash_none(cod),
            "msgErro": dash_none(msg),
            "nivel": dash_none(nivel),
            "execSefin": mark_bool(sefin_exec),
            "sefinDecisao": dash_none(sefin_dec),
            "execAdn": mark_bool(adn_exec),
            "adnDecisao": dash_none(adn_dec),
            "obs": dash_none(obs),
            "fonte": f"{src}!{sheet}:r{r}",
        })
    return out


# RN_RECEPCAO_DPS: # | REGRA(1) | (2) | APLIC(3) | EFEITO(4) | COD(5) | MSG(6) | NOTAS(7)
def extract_regras_recepcao(path, sheet, src):
    mr, mc, grid = load_grid(path, sheet)
    out = []
    for r in range(2, mr + 1):  # row 1 = header
        g = lambda c: grid[r][c] if c <= mc else None
        seq, regra, aplic, efeito, cod, msg, notas = g(1), g(2), g(4), g(5), g(6), g(7), g(8)
        if regra is None:
            continue
        has_code = cod not in (None, "-")
        out.append({
            "seq": to_int(seq),
            "campo": None,
            "regra": regra,
            "aplic": dash_none(aplic),
            "efeito": dash_none(efeito),
            "codErro": dash_none(cod),
            "msgErro": dash_none(msg),
            "secao": None if has_code else regra,  # rows w/o code are section titles
            "obs": dash_none(notas),
            "fonte": f"{src}!{sheet}:r{r}",
        })
    return out


# --------------------------------------------------------------------------- #
# TIPO EVENTOS DE NFSe  (event catalog)
# --------------------------------------------------------------------------- #
def extract_tipos(path, sheet, src):
    mr, mc, grid = load_grid(path, sheet)
    out = []
    for r in range(2, mr + 1):  # row 1 = header
        g = lambda c: grid[r][c] if c <= mc else None
        codigo_raw = g(4)
        codigo = re.sub(r"\s+", "", codigo_raw) if codigo_raw else None
        # stop at the legend row (no numeric event code)
        if not codigo or not re.fullmatch(r"\d{6}", codigo):
            continue
        # the name may span two merged columns; dedupe identical halves
        parts, seen = [flat(g(2)), flat(g(3))], []
        for p in parts:
            if p and p not in seen:
                seen.append(p)
        nome = " - ".join(seen)
        out.append({
            "codigo": codigo,
            "nome": nome,
            "grupoEnvelope": flat(g(5)),  # "NFS-e" etc.
            "categoria": flat(g(6)),      # "1 - Cancelamentos" / "2 - Manifestações" / "3 - Ofícios"
            "autor": flat(g(7)),          # Emite / MEmis / MIncid ...
            "assinaturaObrigatoria": tri_bool(g(8)),
            "ambienteReceptor": flat(g(9)),  # "1 - Sistema que gerou..." / "2 - ADN"
            "precisaExistirNoAdn": tri_bool(g(11)),
            "eventoUnico": tri_bool(g(12)),
            "visibilidade": flat(g(13)),  # "EM / NE / CP / AT"
            "fonte": f"{src}!{sheet}:r{r}",
        })
    return out


# --------------------------------------------------------------------------- #
# RN EVENTOSxEVENTOS  (the event state matrix)
# rows 5-29 = pre-existing state; cols 4-17 = requestable events (block "pedido");
# cols 18-22 = Bloqueio por Ofício / cols 23-27 = Desbloqueio por Ofício,
# each conditioned on an accompanying cancellation sub-type (row 4 label).
# --------------------------------------------------------------------------- #
def extract_sequencia(path, sheet, src, catalog):
    mr, mc, grid = load_grid(path, sheet)
    name2code = {norm(t["nome"]): t["codigo"] for t in catalog}
    # also index by the requestable short names (catalog "nome" already combined)

    def code_for(label):
        n = norm(label)
        if n in name2code:
            return name2code[n]
        for k, v in name2code.items():
            if k and (k in n or n in k):
                return v
        return None

    cells = []
    unmatched_cols = []
    # column metadata (grid is 1-based: requestable events live in cols 5..18,
    # bloqueio por ofício in 19..23, desbloqueio por ofício in 24..28)
    col_meta = {}
    for c in range(5, 19):  # pedido block (14 requestable events)
        label = grid[3][c] if c <= mc else None
        col_meta[c] = {"bloco": "pedido", "requisitado": flat(label),
                       "requisitadoCodigo": code_for(label), "condicao": None}
        if label and not col_meta[c]["requisitadoCodigo"]:
            unmatched_cols.append((c, flat(label)))
    for c in range(19, 24):  # bloqueio por oficio block
        col_meta[c] = {"bloco": "bloqueio-oficio", "requisitado": "Bloqueio de NFS-e por Ofício",
                       "requisitadoCodigo": "305102", "condicao": flat(grid[4][c]) if c <= mc else None}
    for c in range(24, 29):  # desbloqueio por oficio block
        col_meta[c] = {"bloco": "desbloqueio-oficio", "requisitado": "Desbloqueio de NFS-e por Ofício",
                       "requisitadoCodigo": "305103", "condicao": flat(grid[4][c]) if c <= mc else None}

    for r in range(5, mr + 1):
        g = lambda c: grid[r][c] if c <= mc else None
        seq = to_int(g(1))
        pre_grupo = flat(g(3))   # filled from merge for oficio sub-rows
        pre_sub = flat(g(4))     # "Ev. de Cancelamento de NFS-e ..." for oficio rows
        pre = pre_grupo
        if pre is None:
            continue
        for c, meta in col_meta.items():
            marker = g(c)
            if marker is None:
                continue
            cells.append({
                "preExistenteSeq": seq,
                "preExistente": pre,
                "preExistenteSub": pre_sub if pre_sub != pre else None,
                "bloco": meta["bloco"],
                "requisitado": meta["requisitado"],
                "requisitadoCodigo": meta["requisitadoCodigo"],
                "condicaoOficio": meta["condicao"],
                "marcador": marker,
                "permitido": permit(marker),
                "fonte": f"{src}!{sheet}:r{r}c{c}",
            })
    if unmatched_cols:
        print(f"  [warn] sequencia: unmatched requested-event columns -> {unmatched_cols}", file=sys.stderr)
    return cells


# --------------------------------------------------------------------------- #
# Anexo C  IndOp
# Art(0) | Tipo op(1) | Local op(2) | Caracteristica(3) | grupo(4) | seq(5) | codIndOp(6) | local(7) | campo(8)
# --------------------------------------------------------------------------- #
def extract_indop(path, sheet, src):
    mr, mc, grid = load_grid(path, sheet)
    out = []
    for r in range(2, mr + 1):  # row 1 = header
        g = lambda c: grid[r][c] if c <= mc else None
        codigo = g(7)
        if not codigo or not re.search(r"\d", str(codigo)):
            continue
        out.append({
            "codIndOp": str(codigo).strip(),
            "artigo": g(1),
            "tipoOperacao": g(2),
            "localOperacao": g(3),
            "caracteristicaFornecimento": g(4),
            "grupo": g(5),
            "seq": g(6),
            "localFornecimentoIdentificar": g(8),
            "campoLeiaute": g(9),
            "fonte": f"{src}!{sheet}:r{r}",
        })
    return out


# helpers -------------------------------------------------------------------- #
def to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return v


def dash_none(v):
    return None if v in (None, "-") else v


def mark_bool(v):
    """V -> True; '-'/None -> False; anything else kept verbatim."""
    if v is None or v == "-":
        return False
    if v.strip().upper() == "V":
        return True
    return v


def tri_bool(v):
    if v is None or v == "-":
        return None
    s = norm(v)
    if s.startswith("sim"):
        return True
    if s.startswith("nao"):
        return False
    return v


def permit(marker):
    m = marker.strip().upper()
    if m == "V":
        return True
    if m == "X":
        return False
    return "condicional"  # "X/V"


# --------------------------------------------------------------------------- #
def main():
    os.makedirs(RAW, exist_ok=True)
    for p in (A1, A2, AC):
        if not os.path.exists(p):
            sys.exit(f"missing source: {p}")

    summary = {}

    # raw dumps (lossless)
    raw_specs = [
        (A1, "LEIAUTE DPS_NFS-e ", "anexoI.leiaute-dps-nfse"),
        (A1, "RN DPS_NFS-e", "anexoI.rn-dps-nfse"),
        (A1, "RN_RECEPCAO_DPS", "anexoI.rn-recepcao-dps"),
        (A1, "MUN.INCID_INFO.SERV.", "anexoI.mun-incid-info-serv"),
        (A1, "EXPORTACAO_EMISSÃO_NFS-e", "anexoI.exportacao-emissao"),
        (A2, "TIPO EVENTOS DE NFSe", "anexoII.tipo-eventos"),
        (A2, "LEIAUTE EVENTO_PED.REG.EVENTO", "anexoII.leiaute-evento"),
        (A2, "RN EVENTO_PED.REG.EVENTO", "anexoII.rn-evento"),
        (A2, "RN EVENTOSxEVENTOS", "anexoII.rn-eventos-x-eventos"),
        (AC, "IndOp", "anexoC.indop"),
    ]
    for path, sheet, slug in raw_specs:
        dump_raw(path, sheet, slug)

    # clean sidecars
    campos_dps = extract_campos(A1, "LEIAUTE DPS_NFS-e ", "AnexoI")
    write_json(os.path.join(OUT, "emissao.campos.json"), campos_dps)
    summary["emissao.campos"] = len(campos_dps)

    regras_dps = extract_regras_rn(A1, "RN DPS_NFS-e", "AnexoI")
    regras_recep = extract_regras_recepcao(A1, "RN_RECEPCAO_DPS", "AnexoI")
    write_json(os.path.join(OUT, "emissao.regras.json"),
               {"recepcao": regras_recep, "negocio": regras_dps})
    summary["emissao.regras.recepcao"] = len(regras_recep)
    summary["emissao.regras.negocio"] = len(regras_dps)

    tipos = extract_tipos(A2, "TIPO EVENTOS DE NFSe", "AnexoII")
    write_json(os.path.join(OUT, "eventos.tipos.json"), tipos)
    summary["eventos.tipos"] = len(tipos)

    campos_evt = extract_campos(A2, "LEIAUTE EVENTO_PED.REG.EVENTO", "AnexoII")
    write_json(os.path.join(OUT, "eventos.campos.json"), campos_evt)
    summary["eventos.campos"] = len(campos_evt)

    regras_evt = extract_regras_rn(A2, "RN EVENTO_PED.REG.EVENTO", "AnexoII")
    write_json(os.path.join(OUT, "eventos.regras.json"), regras_evt)
    summary["eventos.regras"] = len(regras_evt)

    seq = extract_sequencia(A2, "RN EVENTOSxEVENTOS", "AnexoII", tipos)
    write_json(os.path.join(OUT, "eventos.sequencia.json"), seq)
    summary["eventos.sequencia.cells"] = len(seq)

    indop = extract_indop(AC, "IndOp", "AnexoC")
    write_json(os.path.join(OUT, "tributos.indop.json"), indop)
    summary["tributos.indop"] = len(indop)

    print("== extract-ruleset: clean sidecar counts ==")
    for k, v in summary.items():
        print(f"  {k:32} {v}")
    print(f"== raw dumps: {len(raw_specs)} sheets -> {RAW} ==")


if __name__ == "__main__":
    main()
